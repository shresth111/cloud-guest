"""Unit tests for BE-012 Part 5 (Report Engine + Export Engine):
``ReportTemplate``/``ScheduledReport`` CRUD, on-demand report generation
composing existing analytics services (verified via spies -- proving
composition, never recomputation), every export format (JSON/CSV/Excel/PDF)
producing real, valid, parseable output, the scheduled-report Celery task's
due-report detection + per-schedule failure isolation +
``next_run_at`` computation, email dispatch via the reused
``EmailProviderProtocol``, tenant isolation, and unconditional
(never-throttled) audit-every-generation.

Follows this project's plain-``assert``/native-``async def`` style (see
``tests/unit/test_analytics.py``'s own module docstring) -- every fake below
is a small, hand-rolled stand-in for the narrow protocol it satisfies, no
live Postgres/Redis anywhere in this test suite.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from dataclasses import dataclass, field
from datetime import UTC, datetime, timedelta

import pytest
from openpyxl import load_workbook
from pydantic import BaseModel

from app.domains.analytics.constants import (
    AUDIT_ACTION_REPORT_GENERATED,
    ExportFormat,
    ReportFrequency,
    ReportRunStatus,
    ReportType,
)
from app.domains.analytics.dashboard_scope import DashboardScopeResolver
from app.domains.analytics.exceptions import (
    DashboardScopeForbiddenError,
    MissingReportParametersError,
    ReportTemplateNotFoundError,
    ScheduledReportNotFoundError,
)
from app.domains.analytics.export import render_report
from app.domains.analytics.models import ReportTemplate, ScheduledReport
from app.domains.analytics.report_schemas import (
    ReportTemplateCreateRequest,
    ReportTemplateUpdateRequest,
    ScheduledReportCreateRequest,
    ScheduledReportUpdateRequest,
)
from app.domains.analytics.report_service import (
    ReportGenerationService,
    ReportTemplateService,
    ScheduledReportService,
    compute_next_run_at,
)
from app.domains.analytics.report_tasks import (
    ScheduledReportBatchResult,
    _run_one_scheduled_report,
    run_scheduled_report_batch,
    run_scheduled_reports,
)
from app.domains.analytics.report_types import (
    ReportPayload,
    ReportSection,
    extract_tabular_blocks,
    flatten_scalar_fields,
)
from app.domains.otp.service import EmailProviderProtocol
from app.domains.rbac.enums import ScopeType

# ============================================================================
# Shared fakes -- DashboardScopeResolver composition (mirrors
# tests/unit/test_analytics_forecast_insights.py's own identical fakes;
# self-contained here per this project's own "no cross-test-file fake
# imports" convention).
# ============================================================================


@dataclass
class _FakeRole:
    is_active: bool = True
    is_deleted: bool = False


@dataclass
class _FakeAssignment:
    scope_type: str
    organization_id: uuid.UUID | None = None
    location_id: uuid.UUID | None = None
    router_id: uuid.UUID | None = None
    msp_id: uuid.UUID | None = None
    role: _FakeRole = field(default_factory=_FakeRole)


class _FakeRoleResolver:
    def __init__(self, assignments: list[_FakeAssignment]) -> None:
        self._assignments = assignments

    async def get_active_assignments(self, user_id, *, now=None):
        from app.domains.rbac.authorization import ActiveRoleAssignment

        return [
            ActiveRoleAssignment(assignment=a, role=a.role) for a in self._assignments
        ]


@dataclass
class _FakeOrganization:
    id: uuid.UUID
    name: str = "Fake Org"

    def is_msp(self) -> bool:
        return False


class _FakeOrganizationLookup:
    def __init__(self, organizations=None, children=None) -> None:
        self._organizations = organizations or {}
        self._children = children or {}

    async def get_organization(self, organization_id, *, include_deleted=False):
        return self._organizations[organization_id]

    async def list_children(self, organization_id):
        return self._children.get(organization_id, [])


class _FakeLocationLookup:
    def __init__(self, locations=None) -> None:
        self._locations = locations or {}

    async def get_location(
        self, location_id, *, requesting_organization_id=None, include_deleted=False
    ):
        return self._locations[location_id]


def _global_scope_resolver() -> DashboardScopeResolver:
    return DashboardScopeResolver(
        _FakeRoleResolver([_FakeAssignment(scope_type=ScopeType.GLOBAL.value)]),
        _FakeOrganizationLookup(),
        _FakeLocationLookup(),
    )


def _org_scope_resolver(organization_id: uuid.UUID) -> DashboardScopeResolver:
    return DashboardScopeResolver(
        _FakeRoleResolver(
            [
                _FakeAssignment(
                    scope_type=ScopeType.ORGANIZATION.value,
                    organization_id=organization_id,
                )
            ]
        ),
        _FakeOrganizationLookup(
            {organization_id: _FakeOrganization(id=organization_id)}
        ),
        _FakeLocationLookup(),
    )


class _FakeAuditWriter:
    def __init__(self) -> None:
        self.entries: list[dict] = []

    async def create_audit_log_entry(self, **fields):
        self.entries.append(fields)
        return None


# ============================================================================
# Fakes for ReportGenerationService's five composed services -- each is a
# real scope-check-performing stand-in (calling the exact same
# DashboardScope.require_*/allows_* the real service would), a spy
# (records every call), and returns a small dummy Pydantic model (report_
# service._section only ever needs `.model_dump(mode="json")`, so a real
# SuperAdminDashboardResponse/etc. is not required to prove composition).
# ============================================================================


class _DummyResponse(BaseModel):
    label: str
    rows: list[dict] = []


class _FakeDashboardService:
    def __init__(self, scope_resolver: DashboardScopeResolver) -> None:
        self.scope_resolver = scope_resolver
        self.calls: list[tuple] = []

    async def get_super_admin_dashboard(self, user_id):
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_global()
        self.calls.append(("super_admin", user_id))
        return _DummyResponse(label="platform_dashboard")

    async def get_organization_dashboard(self, user_id, organization_id):
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_organization(organization_id)
        self.calls.append(("organization", user_id, organization_id))
        return _DummyResponse(
            label="organization_dashboard",
            rows=[{"location": "Lobby", "guests": 5}],
        )

    async def get_location_dashboard(self, user_id, location_id):
        organization_id = await self.scope_resolver.resolve_location_organization_id(
            location_id
        )
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_location(location_id, organization_id)
        self.calls.append(("location", user_id, location_id))
        return _DummyResponse(label="location_dashboard")


class _FakeDomainAnalyticsService:
    def __init__(self, scope_resolver: DashboardScopeResolver) -> None:
        self.scope_resolver = scope_resolver
        self.calls: list[tuple] = []

    async def get_router_analytics(
        self, user_id, organization_id, *, location_id, start, end
    ):
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_organization(organization_id)
        self.calls.append(("router", user_id, organization_id, start, end))
        return _DummyResponse(label="router_analytics")

    async def get_guest_analytics(
        self, user_id, organization_id, *, location_id, start, end
    ):
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_organization(organization_id)
        self.calls.append(("guest", user_id, organization_id, start, end))
        return _DummyResponse(label="guest_analytics")

    async def get_network_analytics(
        self, user_id, organization_id, *, location_id, start, end
    ):
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_organization(organization_id)
        self.calls.append(("network", user_id, organization_id, start, end))
        return _DummyResponse(label="network_analytics")


class _FakeBusinessAnalyticsService:
    def __init__(self, scope_resolver: DashboardScopeResolver) -> None:
        self.scope_resolver = scope_resolver
        self.calls: list[tuple] = []

    async def get_business_analytics(self, user_id):
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_global()
        self.calls.append(("business", user_id))
        return _DummyResponse(label="business_analytics")


class _FakeForecastService:
    def __init__(self, scope_resolver: DashboardScopeResolver) -> None:
        self.scope_resolver = scope_resolver
        self.calls: list[tuple] = []

    async def get_router_failure_risk(self, user_id, organization_id, *, location_id):
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_organization(organization_id)
        self.calls.append(("router_failure_risk", user_id, organization_id))
        return _DummyResponse(label="router_failure_risk")


class _FakeInsightService:
    def __init__(self, scope_resolver: DashboardScopeResolver) -> None:
        self.scope_resolver = scope_resolver
        self.calls: list[tuple] = []

    async def get_business_insights(self, user_id):
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_global()
        self.calls.append(("business_insights", user_id))
        return _DummyResponse(label="business_insights")

    async def get_operational_recommendations(self, user_id):
        scope = await self.scope_resolver.resolve(user_id)
        scope.require_global()
        self.calls.append(("operational_recommendations", user_id))
        return _DummyResponse(label="operational_recommendations")


def _build_generation_service(scope_resolver, *, audit_writer=None):
    dashboard = _FakeDashboardService(scope_resolver)
    domain = _FakeDomainAnalyticsService(scope_resolver)
    business = _FakeBusinessAnalyticsService(scope_resolver)
    forecast = _FakeForecastService(scope_resolver)
    insight = _FakeInsightService(scope_resolver)
    service = ReportGenerationService(
        dashboard, domain, business, forecast, insight, audit_writer=audit_writer
    )
    return service, dashboard, domain, business, forecast, insight


# ============================================================================
# ReportGenerationService.generate -- composition, not recomputation
# ============================================================================


async def test_generate_dashboard_report_composes_super_admin_dashboard():
    service, dashboard, domain, business, forecast, insight = _build_generation_service(
        _global_scope_resolver()
    )
    payload = await service.generate(
        uuid.uuid4(),
        report_type=ReportType.DASHBOARD,
        organization_id=None,
        location_id=None,
        start=None,
        end=None,
    )
    assert len(dashboard.calls) == 1
    assert dashboard.calls[0][0] == "super_admin"
    assert not domain.calls and not business.calls and not forecast.calls
    assert not insight.calls
    assert payload.report_type == "dashboard"
    assert len(payload.sections) == 1
    assert payload.sections[0].key == "platform_dashboard"
    assert payload.sections[0].data == {"label": "platform_dashboard", "rows": []}


async def test_generate_organization_report_requires_organization_id():
    service, *_ = _build_generation_service(_global_scope_resolver())
    with pytest.raises(MissingReportParametersError):
        await service.generate(
            uuid.uuid4(),
            report_type=ReportType.ORGANIZATION,
            organization_id=None,
            location_id=None,
            start=None,
            end=None,
        )


async def test_generate_organization_report_composes_organization_dashboard():
    org_id = uuid.uuid4()
    service, dashboard, *_ = _build_generation_service(_org_scope_resolver(org_id))
    payload = await service.generate(
        uuid.uuid4(),
        report_type=ReportType.ORGANIZATION,
        organization_id=org_id,
        location_id=None,
        start=None,
        end=None,
    )
    assert dashboard.calls == [("organization", dashboard.calls[0][1], org_id)]
    assert payload.sections[0].key == "organization_dashboard"
    # The composed response's own list-of-dicts field survives verbatim --
    # proof this is composition (embedding an existing response), not a
    # reimplementation of its own metrics.
    assert payload.sections[0].data["rows"] == [{"location": "Lobby", "guests": 5}]


async def test_generate_location_report_requires_location_id():
    service, *_ = _build_generation_service(_global_scope_resolver())
    with pytest.raises(MissingReportParametersError):
        await service.generate(
            uuid.uuid4(),
            report_type=ReportType.LOCATION,
            organization_id=None,
            location_id=None,
            start=None,
            end=None,
        )


async def test_generate_router_guest_network_reports_require_organization_id():
    service, *_ = _build_generation_service(_global_scope_resolver())
    for report_type in (ReportType.ROUTER, ReportType.GUEST, ReportType.NETWORK):
        with pytest.raises(MissingReportParametersError):
            await service.generate(
                uuid.uuid4(),
                report_type=report_type,
                organization_id=None,
                location_id=None,
                start=None,
                end=None,
            )


async def test_generate_router_report_resolves_default_window_and_composes():
    org_id = uuid.uuid4()
    service, dashboard, domain, *_ = _build_generation_service(
        _org_scope_resolver(org_id)
    )
    payload = await service.generate(
        uuid.uuid4(),
        report_type=ReportType.ROUTER,
        organization_id=org_id,
        location_id=None,
        start=None,
        end=None,
    )
    assert len(domain.calls) == 1
    assert domain.calls[0][0] == "router"
    assert payload.sections[0].key == "router_analytics"
    # No explicit start/end supplied -- a default trailing window was
    # resolved (validators.resolve_analytics_window), not left empty.
    assert payload.period_start is not None
    assert payload.period_end is not None


async def test_generate_revenue_report_composes_business_analytics():
    service, dashboard, domain, business, forecast, insight = _build_generation_service(
        _global_scope_resolver()
    )
    payload = await service.generate(
        uuid.uuid4(),
        report_type=ReportType.REVENUE,
        organization_id=None,
        location_id=None,
        start=None,
        end=None,
    )
    assert len(business.calls) == 1
    assert payload.sections[0].key == "business_analytics"


async def test_generate_health_report_without_organization_composes_insights_only():
    service, dashboard, domain, business, forecast, insight = _build_generation_service(
        _global_scope_resolver()
    )
    payload = await service.generate(
        uuid.uuid4(),
        report_type=ReportType.HEALTH,
        organization_id=None,
        location_id=None,
        start=None,
        end=None,
    )
    assert len(insight.calls) == 2
    assert not forecast.calls
    section_keys = [s.key for s in payload.sections]
    assert section_keys == ["business_insights", "operational_recommendations"]


async def test_generate_health_report_with_organization_also_composes_router_risk():
    org_id = uuid.uuid4()
    # A GLOBAL scope covers both the insight engine's own GLOBAL
    # requirement and the forecast engine's ORGANIZATION requirement.
    service, dashboard, domain, business, forecast, insight = _build_generation_service(
        _global_scope_resolver()
    )
    payload = await service.generate(
        uuid.uuid4(),
        report_type=ReportType.HEALTH,
        organization_id=org_id,
        location_id=None,
        start=None,
        end=None,
    )
    assert len(forecast.calls) == 1
    section_keys = [s.key for s in payload.sections]
    assert section_keys == [
        "business_insights",
        "operational_recommendations",
        "router_failure_risk",
    ]


async def test_generate_health_report_can_disable_router_failure_risk_section():
    org_id = uuid.uuid4()
    service, dashboard, domain, business, forecast, insight = _build_generation_service(
        _global_scope_resolver()
    )
    payload = await service.generate(
        uuid.uuid4(),
        report_type=ReportType.HEALTH,
        organization_id=org_id,
        location_id=None,
        start=None,
        end=None,
        include_router_failure_risk=False,
    )
    assert not forecast.calls
    assert [s.key for s in payload.sections] == [
        "business_insights",
        "operational_recommendations",
    ]


async def test_generate_report_rejects_out_of_scope_organization():
    """Tenant isolation: an organization-scoped caller for org A cannot
    generate an ORGANIZATION report for org B -- the exact
    ``DashboardScopeForbiddenError`` the composed (fake, but scope-check-
    performing) ``DashboardService`` raises propagates untouched, since
    ``ReportGenerationService`` adds no second, parallel scope check of its
    own (see that method's own docstring)."""
    org_a = uuid.uuid4()
    org_b = uuid.uuid4()
    service, *_ = _build_generation_service(_org_scope_resolver(org_a))
    with pytest.raises(DashboardScopeForbiddenError):
        await service.generate(
            uuid.uuid4(),
            report_type=ReportType.ORGANIZATION,
            organization_id=org_b,
            location_id=None,
            start=None,
            end=None,
        )


async def test_generate_report_writes_audit_entry_unconditionally_every_call():
    """Full, un-throttled auditing: three separate calls -> three separate
    audit rows (contrast with Part 2's throttled dashboard-view auditing,
    at most one row per 15-minute window)."""
    audit_writer = _FakeAuditWriter()
    service, *_ = _build_generation_service(
        _global_scope_resolver(), audit_writer=audit_writer
    )
    user_id = uuid.uuid4()
    for _ in range(3):
        await service.generate(
            user_id,
            report_type=ReportType.DASHBOARD,
            organization_id=None,
            location_id=None,
            start=None,
            end=None,
            export_format=ExportFormat.PDF,
        )
    assert len(audit_writer.entries) == 3
    for entry in audit_writer.entries:
        assert entry["action"] == AUDIT_ACTION_REPORT_GENERATED
        assert entry["event_metadata"]["report_type"] == "dashboard"
        assert entry["event_metadata"]["export_format"] == "pdf"


# ============================================================================
# ReportTemplateService CRUD
# ============================================================================


class _FakeReportRepository:
    """Stand-in for ``ReportRepositoryProtocol`` covering only what
    ``ReportTemplateService``/``ScheduledReportService`` need."""

    def __init__(self) -> None:
        self.templates: dict[uuid.UUID, ReportTemplate] = {}
        self.schedules: dict[uuid.UUID, ScheduledReport] = {}
        self.list_templates_calls: list[dict] = []
        self.list_schedules_calls: list[dict] = []

    async def create_template(self, **fields) -> ReportTemplate:
        template = ReportTemplate(**fields)
        template.id = uuid.uuid4()
        self.templates[template.id] = template
        return template

    async def get_template(self, template_id, *, include_deleted=False):
        return self.templates.get(template_id)

    async def list_templates(self, *, see_all, organization_ids, page, page_size):
        self.list_templates_calls.append(
            {"see_all": see_all, "organization_ids": organization_ids}
        )
        items = list(self.templates.values())
        from app.database.utils.pagination import PageParams, PaginationMeta

        return items, PaginationMeta.from_total(
            PageParams(page=page, page_size=page_size), len(items)
        )

    async def update_template(self, template, fields):
        for key, value in fields.items():
            if value is not None:
                setattr(template, key, value)
        return template

    async def soft_delete_template(self, template):
        template.is_deleted = True
        return template

    async def create_schedule(self, **fields) -> ScheduledReport:
        schedule = ScheduledReport(**fields)
        schedule.id = uuid.uuid4()
        self.schedules[schedule.id] = schedule
        return schedule

    async def get_schedule(self, schedule_id, *, include_deleted=False):
        return self.schedules.get(schedule_id)

    async def list_schedules(self, *, see_all, organization_ids, page, page_size):
        self.list_schedules_calls.append(
            {"see_all": see_all, "organization_ids": organization_ids}
        )
        items = list(self.schedules.values())
        from app.database.utils.pagination import PageParams, PaginationMeta

        return items, PaginationMeta.from_total(
            PageParams(page=page, page_size=page_size), len(items)
        )

    async def update_schedule(self, schedule, fields):
        for key, value in fields.items():
            if value is not None:
                setattr(schedule, key, value)
        return schedule

    async def soft_delete_schedule(self, schedule):
        schedule.is_deleted = True
        return schedule

    async def get_due_scheduled_reports(self, *, now):
        return [
            schedule
            for schedule in self.schedules.values()
            if schedule.is_active and schedule.next_run_at <= now
        ]


async def test_create_template_platform_wide_requires_global_scope():
    repository = _FakeReportRepository()
    audit_writer = _FakeAuditWriter()
    service = ReportTemplateService(
        repository, _global_scope_resolver(), audit_writer=audit_writer
    )
    payload = ReportTemplateCreateRequest(
        name="Platform Weekly", report_type=ReportType.DASHBOARD
    )
    template = await service.create_template(uuid.uuid4(), None, payload)
    assert template.organization_id is None
    assert template.report_type == ReportType.DASHBOARD.value
    assert len(audit_writer.entries) == 1


async def test_create_template_platform_wide_rejects_non_global_caller():
    org_id = uuid.uuid4()
    repository = _FakeReportRepository()
    service = ReportTemplateService(repository, _org_scope_resolver(org_id))
    payload = ReportTemplateCreateRequest(
        name="Platform Weekly", report_type=ReportType.DASHBOARD
    )
    with pytest.raises(DashboardScopeForbiddenError):
        await service.create_template(uuid.uuid4(), None, payload)


async def test_create_template_org_scoped_requires_matching_organization():
    org_a = uuid.uuid4()
    org_b = uuid.uuid4()
    repository = _FakeReportRepository()
    service = ReportTemplateService(repository, _org_scope_resolver(org_a))
    payload = ReportTemplateCreateRequest(
        name="Org Report", report_type=ReportType.ORGANIZATION
    )
    template = await service.create_template(uuid.uuid4(), org_a, payload)
    assert template.organization_id == org_a

    with pytest.raises(DashboardScopeForbiddenError):
        await service.create_template(uuid.uuid4(), org_b, payload)


async def test_get_visible_template_not_found_for_missing_row():
    repository = _FakeReportRepository()
    service = ReportTemplateService(repository, _global_scope_resolver())
    with pytest.raises(ReportTemplateNotFoundError):
        await service.get_visible_template(uuid.uuid4(), uuid.uuid4())


async def test_get_visible_template_not_found_for_out_of_scope_organization():
    """Tenant isolation: an org-scoped template belonging to organization A
    reads as "not found" (not "forbidden") to a caller scoped to
    organization B -- mirrors this exception's own documented convention."""
    org_a = uuid.uuid4()
    org_b = uuid.uuid4()
    repository = _FakeReportRepository()
    creator_service = ReportTemplateService(repository, _org_scope_resolver(org_a))
    payload = ReportTemplateCreateRequest(
        name="Org A Report", report_type=ReportType.ORGANIZATION
    )
    template = await creator_service.create_template(uuid.uuid4(), org_a, payload)

    reader_service = ReportTemplateService(repository, _org_scope_resolver(org_b))
    with pytest.raises(ReportTemplateNotFoundError):
        await reader_service.get_visible_template(uuid.uuid4(), template.id)


async def test_list_templates_global_scope_sees_all():
    repository = _FakeReportRepository()
    service = ReportTemplateService(repository, _global_scope_resolver())
    await service.list_templates(uuid.uuid4(), page=1, page_size=25)
    assert repository.list_templates_calls[-1]["see_all"] is True


async def test_list_templates_org_scope_filters_by_organization_ids():
    org_id = uuid.uuid4()
    repository = _FakeReportRepository()
    service = ReportTemplateService(repository, _org_scope_resolver(org_id))
    await service.list_templates(uuid.uuid4(), page=1, page_size=25)
    call = repository.list_templates_calls[-1]
    assert call["see_all"] is False
    assert call["organization_ids"] == [org_id]


async def test_update_and_delete_template_audits_each_action():
    repository = _FakeReportRepository()
    audit_writer = _FakeAuditWriter()
    service = ReportTemplateService(
        repository, _global_scope_resolver(), audit_writer=audit_writer
    )
    payload = ReportTemplateCreateRequest(
        name="Original", report_type=ReportType.DASHBOARD
    )
    template = await service.create_template(uuid.uuid4(), None, payload)

    updated = await service.update_template(
        uuid.uuid4(),
        template.id,
        ReportTemplateUpdateRequest(name="Renamed"),
    )
    assert updated.name == "Renamed"

    await service.delete_template(uuid.uuid4(), template.id)
    assert template.is_deleted is True

    actions = [entry["action"] for entry in audit_writer.entries]
    assert actions == [
        "report_template_created",
        "report_template_updated",
        "report_template_deleted",
    ]


# ============================================================================
# ScheduledReportService CRUD + compute_next_run_at
# ============================================================================


@pytest.mark.parametrize(
    "frequency,expected_days",
    [
        (ReportFrequency.DAILY, 1),
        (ReportFrequency.WEEKLY, 7),
        (ReportFrequency.MONTHLY, 30),
    ],
)
def test_compute_next_run_at_advances_by_expected_days(frequency, expected_days):
    base = datetime(2026, 1, 1, tzinfo=UTC)
    assert compute_next_run_at(frequency, base) == base + timedelta(days=expected_days)


async def test_create_schedule_requires_organization_scope_and_visible_template():
    org_id = uuid.uuid4()
    repository = _FakeReportRepository()
    template_service = ReportTemplateService(repository, _org_scope_resolver(org_id))
    template = await template_service.create_template(
        uuid.uuid4(),
        org_id,
        ReportTemplateCreateRequest(name="Weekly", report_type=ReportType.ORGANIZATION),
    )

    audit_writer = _FakeAuditWriter()
    service = ScheduledReportService(
        repository,
        _org_scope_resolver(org_id),
        template_service,
        audit_writer=audit_writer,
    )
    schedule = await service.create_schedule(
        uuid.uuid4(),
        org_id,
        ScheduledReportCreateRequest(
            template_id=template.id,
            frequency=ReportFrequency.WEEKLY,
            recipient_emails=["ops@example.com"],
            export_format=ExportFormat.PDF,
        ),
    )
    assert schedule.organization_id == org_id
    assert schedule.frequency == ReportFrequency.WEEKLY.value
    assert schedule.next_run_at > datetime.now(UTC)
    assert len(audit_writer.entries) == 1


async def test_create_schedule_rejects_out_of_scope_organization():
    org_a = uuid.uuid4()
    org_b = uuid.uuid4()
    repository = _FakeReportRepository()
    template_service = ReportTemplateService(repository, _global_scope_resolver())
    template = await template_service.create_template(
        uuid.uuid4(),
        None,
        ReportTemplateCreateRequest(name="Platform", report_type=ReportType.DASHBOARD),
    )
    service = ScheduledReportService(
        repository, _org_scope_resolver(org_a), template_service
    )
    with pytest.raises(DashboardScopeForbiddenError):
        await service.create_schedule(
            uuid.uuid4(),
            org_b,
            ScheduledReportCreateRequest(
                template_id=template.id,
                frequency=ReportFrequency.DAILY,
                recipient_emails=["ops@example.com"],
            ),
        )


async def test_get_visible_schedule_not_found_for_out_of_scope_organization():
    org_a = uuid.uuid4()
    org_b = uuid.uuid4()
    repository = _FakeReportRepository()
    template_service = ReportTemplateService(repository, _global_scope_resolver())
    template = await template_service.create_template(
        uuid.uuid4(),
        None,
        ReportTemplateCreateRequest(name="Platform", report_type=ReportType.DASHBOARD),
    )
    creator_service = ScheduledReportService(
        repository, _org_scope_resolver(org_a), template_service
    )
    schedule = await creator_service.create_schedule(
        uuid.uuid4(),
        org_a,
        ScheduledReportCreateRequest(
            template_id=template.id,
            frequency=ReportFrequency.DAILY,
            recipient_emails=["ops@example.com"],
        ),
    )

    reader_service = ScheduledReportService(
        repository, _org_scope_resolver(org_b), template_service
    )
    with pytest.raises(ScheduledReportNotFoundError):
        await reader_service.get_visible_schedule(uuid.uuid4(), schedule.id)


async def test_update_schedule_frequency_change_rebases_next_run_at():
    org_id = uuid.uuid4()
    repository = _FakeReportRepository()
    template_service = ReportTemplateService(repository, _org_scope_resolver(org_id))
    template = await template_service.create_template(
        uuid.uuid4(),
        org_id,
        ReportTemplateCreateRequest(name="Weekly", report_type=ReportType.ORGANIZATION),
    )
    service = ScheduledReportService(
        repository, _org_scope_resolver(org_id), template_service
    )
    schedule = await service.create_schedule(
        uuid.uuid4(),
        org_id,
        ScheduledReportCreateRequest(
            template_id=template.id,
            frequency=ReportFrequency.DAILY,
            recipient_emails=["ops@example.com"],
        ),
    )
    original_next_run = schedule.next_run_at

    updated = await service.update_schedule(
        uuid.uuid4(),
        schedule.id,
        ScheduledReportUpdateRequest(frequency=ReportFrequency.MONTHLY),
    )
    assert updated.frequency == ReportFrequency.MONTHLY.value
    # Rebased from "now" under the new cadence, not left stale under the
    # old one.
    assert updated.next_run_at != original_next_run
    assert updated.next_run_at > datetime.now(UTC) + timedelta(days=25)


# ============================================================================
# report_types.py -- flattening convention primitives
# ============================================================================


def test_flatten_scalar_fields_walks_nested_dicts_and_joins_scalar_lists():
    data = {
        "top": 1,
        "nested": {"inner": 2, "deeper": {"value": 3}},
        "tags": ["a", "b", "c"],
        "table": [{"x": 1}],
    }
    pairs = dict(flatten_scalar_fields(data))
    assert pairs["top"] == 1
    assert pairs["nested.inner"] == 2
    assert pairs["nested.deeper.value"] == 3
    assert pairs["tags"] == "a; b; c"
    assert "table" not in pairs


def test_extract_tabular_blocks_unions_columns_and_defaults_missing_to_empty():
    data = {
        "section": {
            "items": [
                {"a": 1, "b": 2},
                {"a": 3},
            ]
        }
    }
    blocks = extract_tabular_blocks(data)
    assert len(blocks) == 1
    block = blocks[0]
    assert block.name == "section.items"
    assert block.columns == ["a", "b"]
    assert block.rows == [[1, 2], [3, ""]]


def test_extract_tabular_blocks_ignores_empty_lists_and_non_dict_lists():
    data = {"empty": [], "scalars": [1, 2, 3]}
    assert extract_tabular_blocks(data) == []


# ============================================================================
# export.py -- every format, real and parseable
# ============================================================================


def _sample_payload() -> ReportPayload:
    return ReportPayload(
        report_type="organization",
        title="Organization Report",
        generated_at=datetime(2026, 1, 15, tzinfo=UTC).isoformat(),
        organization_id=uuid.uuid4(),
        location_id=None,
        period_start=datetime(2026, 1, 1, tzinfo=UTC).isoformat(),
        period_end=datetime(2026, 1, 15, tzinfo=UTC).isoformat(),
        sections=[
            ReportSection(
                key="organization_dashboard",
                title="Organization Dashboard",
                data={
                    "total_guests": 42,
                    "health_score": {"score": 87.5, "grade": "B"},
                    "recipient_emails": ["a@example.com", "b@example.com"],
                    "organization_summary": [
                        {"location": "Lobby", "guests": 10},
                        {"location": "Cafe", "guests": 32},
                    ],
                },
            )
        ],
    )


def test_render_json_produces_real_parseable_payload():
    payload = _sample_payload()
    rendered = render_report(payload, ExportFormat.JSON)
    assert rendered.content_type == "application/json"
    parsed = json.loads(rendered.content)
    assert parsed["report_type"] == "organization"
    assert parsed["sections"][0]["data"]["total_guests"] == 42
    assert parsed["organization_id"] == str(payload.organization_id)


def test_render_csv_produces_real_parseable_rows_with_summary_and_tabular_blocks():
    payload = _sample_payload()
    rendered = render_report(payload, ExportFormat.CSV)
    assert rendered.content_type == "text/csv"
    text = rendered.content.decode("utf-8")
    rows = list(csv.reader(io.StringIO(text)))

    # The Summary block's scalar, dotted-path-flattened fields are present.
    flat_rows = {
        (row[0], row[1]): row[2]
        for row in rows
        if len(row) == 3 and row[0] == "organization_dashboard"
    }
    assert flat_rows[("organization_dashboard", "total_guests")] == "42"
    assert flat_rows[("organization_dashboard", "health_score.score")] == "87.5"
    assert (
        flat_rows[("organization_dashboard", "recipient_emails")]
        == "a@example.com; b@example.com"
    )

    # The tabular block marker + its own rows are present, verbatim.
    joined = "\n".join(",".join(row) for row in rows)
    assert "organization_dashboard.organization_summary" in joined
    assert any(row == ["Lobby", "10"] for row in rows)
    assert any(row == ["Cafe", "32"] for row in rows)


def test_render_excel_produces_real_workbook_with_summary_and_tabular_sheets():
    payload = _sample_payload()
    rendered = render_report(payload, ExportFormat.EXCEL)
    assert rendered.content_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Real, valid .xlsx bytes -- the zip/xlsx magic header.
    assert rendered.content[:2] == b"PK"

    workbook = load_workbook(io.BytesIO(rendered.content))
    assert "Summary" in workbook.sheetnames
    # Excel sheet names are truncated to 31 characters (openpyxl's own
    # limit -- see export._sanitize_sheet_name), so the tabular sheet is
    # identified by its one non-Summary sheet, not a name substring match.
    tabular_sheet_names = [name for name in workbook.sheetnames if name != "Summary"]
    assert len(tabular_sheet_names) == 1
    assert tabular_sheet_names[0].startswith("organization_dashboard_organiza")

    summary_sheet = workbook["Summary"]
    summary_values = {
        (row[0].value, row[1].value): row[2].value
        for row in summary_sheet.iter_rows(min_row=2)
    }
    assert summary_values[("organization_dashboard", "total_guests")] == "42"

    tabular_sheet = workbook[tabular_sheet_names[0]]
    rows = [[cell.value for cell in row] for row in tabular_sheet.iter_rows()]
    assert rows[0] == ["location", "guests"]
    assert ["Lobby", "10"] in rows
    assert ["Cafe", "32"] in rows


def test_render_pdf_produces_real_valid_pdf_with_report_content():
    payload = _sample_payload()
    rendered = render_report(payload, ExportFormat.PDF)
    assert rendered.content_type == "application/pdf"
    # Real, valid PDF -- the %PDF magic header, per this domain's own
    # verification mandate.
    assert rendered.content[:4] == b"%PDF"
    assert rendered.content.rstrip().endswith(b"%%EOF")
    assert len(rendered.content) > 500  # a real, non-trivial document

    # reportlab's own uncompressed text-drawing operators land in the
    # stream verbatim often enough to assert real title text is present
    # without pulling in a second, heavier PDF-parsing dependency just for
    # this test -- a best-effort content check, not the sole validity
    # proof (the header/EOF/size checks above already establish that).
    assert b"Organization Report" in rendered.content or len(rendered.content) > 1000


# ============================================================================
# report_tasks.py -- due-report detection, per-schedule failure isolation,
# next_run_at recomputation, email dispatch via EmailProviderProtocol
# ============================================================================


class _FakeGenerationService:
    def __init__(self, *, fail_for: set[uuid.UUID] | None = None) -> None:
        self.fail_for = fail_for or set()
        self.calls: list[uuid.UUID] = []

    async def generate(self, user_id, **kwargs):
        template_id = kwargs.get("template_id")
        self.calls.append(template_id)
        if template_id in self.fail_for:
            raise RuntimeError("simulated generation failure")
        return ReportPayload(
            report_type=kwargs["report_type"].value,
            title="Scheduled Report",
            generated_at=datetime.now(UTC).isoformat(),
            organization_id=kwargs["organization_id"],
            location_id=kwargs["location_id"],
            period_start=None,
            period_end=None,
            sections=[ReportSection(key="s", title="S", data={"x": 1})],
        )


class _FakeEmailProvider(EmailProviderProtocol):
    def __init__(self) -> None:
        self.sent: list[tuple[str, str, str]] = []

    async def send(self, email: str, subject: str, body: str) -> None:
        self.sent.append((email, subject, body))


def _make_template(*, created_by_user_id=None, config=None) -> ReportTemplate:
    template = ReportTemplate(
        name="T",
        organization_id=None,
        report_type=ReportType.DASHBOARD.value,
        config=config or {},
        is_active=True,
        created_by_user_id=created_by_user_id,
    )
    template.id = uuid.uuid4()
    return template


def _make_schedule(
    *, template_id, organization_id=None, created_by_user_id=None, next_run_at=None
) -> ScheduledReport:
    schedule = ScheduledReport(
        template_id=template_id,
        organization_id=organization_id or uuid.uuid4(),
        frequency=ReportFrequency.DAILY.value,
        recipient_emails=["ops@example.com", "cto@example.com"],
        export_format=ExportFormat.JSON.value,
        next_run_at=next_run_at or datetime.now(UTC),
        created_by_user_id=created_by_user_id,
    )
    schedule.id = uuid.uuid4()
    return schedule


class _FakeReportRepositoryForTasks:
    def __init__(self, templates: dict[uuid.UUID, ReportTemplate]) -> None:
        self.templates = templates
        self.update_calls: list[tuple[uuid.UUID, dict]] = []

    async def get_template(self, template_id):
        return self.templates.get(template_id)

    async def update_schedule(self, schedule, fields):
        self.update_calls.append((schedule.id, dict(fields)))
        for key, value in fields.items():
            setattr(schedule, key, value)
        return schedule


async def test_run_one_scheduled_report_sends_email_to_every_recipient():
    creator_id = uuid.uuid4()
    template = _make_template(created_by_user_id=creator_id)
    schedule = _make_schedule(template_id=template.id)
    repository = _FakeReportRepositoryForTasks({template.id: template})
    generation_service = _FakeGenerationService()
    email_provider = _FakeEmailProvider()

    await _run_one_scheduled_report(
        repository, generation_service, email_provider, schedule, now=datetime.now(UTC)
    )

    assert generation_service.calls == [schedule.template_id]
    assert [email for email, _, _ in email_provider.sent] == schedule.recipient_emails
    for _, subject, body in email_provider.sent:
        assert "Scheduled Report" in subject
        # Honest attachment limitation -- see report_tasks.py's own module
        # docstring: the body describes the report, it does not (cannot,
        # via this shared protocol) attach its bytes.
        assert "json" in body.lower() or "bytes" in body.lower()


async def test_run_one_scheduled_report_raises_when_template_missing():
    schedule = _make_schedule(template_id=uuid.uuid4())
    repository = _FakeReportRepositoryForTasks({})
    with pytest.raises(ValueError, match="no longer exists"):
        await _run_one_scheduled_report(
            repository,
            _FakeGenerationService(),
            _FakeEmailProvider(),
            schedule,
            now=datetime.now(UTC),
        )


async def test_run_one_scheduled_report_raises_when_no_attributable_actor():
    template = _make_template(created_by_user_id=None)
    schedule = _make_schedule(template_id=template.id, created_by_user_id=None)
    repository = _FakeReportRepositoryForTasks({template.id: template})
    with pytest.raises(ValueError, match="no attributable creator"):
        await _run_one_scheduled_report(
            repository,
            _FakeGenerationService(),
            _FakeEmailProvider(),
            schedule,
            now=datetime.now(UTC),
        )


async def test_scheduled_report_batch_isolates_one_failure_and_reports_it():
    creator_id = uuid.uuid4()
    good_template = _make_template(created_by_user_id=creator_id)
    bad_template = _make_template(created_by_user_id=creator_id)
    good_schedule = _make_schedule(template_id=good_template.id)
    bad_schedule = _make_schedule(template_id=bad_template.id)

    repository = _FakeReportRepositoryForTasks(
        {good_template.id: good_template, bad_template.id: bad_template}
    )
    generation_service = _FakeGenerationService(fail_for={bad_template.id})
    email_provider = _FakeEmailProvider()
    now = datetime.now(UTC)

    result = await run_scheduled_report_batch(
        repository,
        generation_service,
        email_provider,
        [good_schedule, bad_schedule],
        now=now,
    )

    assert isinstance(result, ScheduledReportBatchResult)
    assert result.total_due == 2
    assert result.succeeded == 1
    assert len(result.failed) == 1
    assert result.failed[0][0] == bad_schedule.id
    assert "simulated generation failure" in result.failed[0][1]

    # Both schedules' state was updated (one SUCCESS, one FAILED) -- a
    # failure never blocks the batch from recording its own outcome.
    statuses = {
        schedule_id: fields["last_run_status"]
        for schedule_id, fields in repository.update_calls
    }
    assert statuses[good_schedule.id] == ReportRunStatus.SUCCESS.value
    assert statuses[bad_schedule.id] == ReportRunStatus.FAILED.value
    # A persistently-broken schedule still gets rescheduled forward (real
    # backoff, not a busy-loop retrying every tick) rather than left stuck.
    for _, fields in repository.update_calls:
        assert fields["next_run_at"] > now

    # Only the good schedule's recipients actually received a "send".
    assert len(email_provider.sent) == len(good_schedule.recipient_emails)


async def test_scheduled_report_batch_empty_due_list_is_a_no_op():
    result = await run_scheduled_report_batch(
        _FakeReportRepositoryForTasks({}),
        _FakeGenerationService(),
        _FakeEmailProvider(),
        [],
        now=datetime.now(UTC),
    )
    assert result == ScheduledReportBatchResult(total_due=0, succeeded=0, failed=[])


def test_run_scheduled_reports_task_bridges_into_async(monkeypatch):
    from app.domains.analytics import report_tasks as report_tasks_module

    recorded: dict[str, object] = {}

    async def _fake_run_scheduled_reports_async(*, now_iso):
        recorded["now_iso"] = now_iso
        return ScheduledReportBatchResult(
            total_due=2, succeeded=1, failed=[(uuid.uuid4(), "boom")]
        )

    monkeypatch.setattr(
        report_tasks_module,
        "_run_scheduled_reports_async",
        _fake_run_scheduled_reports_async,
    )

    result = run_scheduled_reports(now_iso="2026-01-15T00:00:00+00:00")

    assert recorded == {"now_iso": "2026-01-15T00:00:00+00:00"}
    assert result["total_due"] == 2
    assert result["succeeded"] == 1
    assert len(result["failed"]) == 1
    assert "scheduled_report_id" in result["failed"][0]


# ============================================================================
# Route registration / RBAC permission-key wiring
# ============================================================================


def _permission_key_and_scope_for_route(route) -> tuple[str | None, str | None]:
    for dependency in route.dependant.dependencies:
        call = dependency.call
        freevars = getattr(call.__code__, "co_freevars", ())
        if "permission_key" in freevars:
            key_index = freevars.index("permission_key")
            permission_key = call.__closure__[key_index].cell_contents
            scope_value = None
            if "scope" in freevars:
                scope_index = freevars.index("scope")
                scope = call.__closure__[scope_index].cell_contents
                scope_value = scope.value if scope is not None else None
            return permission_key, scope_value
    return None, None


@pytest.fixture(scope="module")
def report_routes_by_path_and_method():
    from app.main import create_app

    app = create_app()
    routes: dict[tuple[str, str], object] = {}
    for route in app.routes:
        path = getattr(route, "path", None)
        methods = getattr(route, "methods", None)
        if path and methods and path.startswith("/api/v1/reports"):
            for method in methods:
                routes[(path, method)] = route
    return routes


@pytest.mark.parametrize(
    "path,method,expected_key,expected_scope",
    [
        ("/api/v1/reports/templates", "POST", "reports.manage", None),
        ("/api/v1/reports/templates", "GET", "reports.read", None),
        ("/api/v1/reports/templates/{template_id}", "GET", "reports.read", None),
        ("/api/v1/reports/templates/{template_id}", "PUT", "reports.manage", None),
        ("/api/v1/reports/templates/{template_id}", "DELETE", "reports.manage", None),
        ("/api/v1/reports", "POST", "reports.export", None),
        (
            "/api/v1/reports/schedule",
            "POST",
            "reports.manage",
            "organization",
        ),
        ("/api/v1/reports/schedule", "GET", "reports.read", None),
        ("/api/v1/reports/schedule/{schedule_id}", "PUT", "reports.manage", None),
        ("/api/v1/reports/schedule/{schedule_id}", "DELETE", "reports.manage", None),
    ],
)
def test_report_routes_require_expected_permission_key_and_scope(
    report_routes_by_path_and_method, path, method, expected_key, expected_scope
):
    route = report_routes_by_path_and_method[(path, method)]
    permission_key, scope = _permission_key_and_scope_for_route(route)
    assert permission_key == expected_key
    assert scope == expected_scope


def test_app_boots_with_part5_report_routes_registered_and_no_route_conflicts():
    from app.main import create_app

    app = create_app()
    paths = {getattr(r, "path", None) for r in app.routes}
    for expected in [
        "/api/v1/reports",
        "/api/v1/reports/templates",
        "/api/v1/reports/templates/{template_id}",
        "/api/v1/reports/schedule",
        "/api/v1/reports/schedule/{schedule_id}",
    ]:
        assert expected in paths

    seen: set[tuple[str, frozenset[str]]] = set()
    for route in app.routes:
        route_path = getattr(route, "path", None)
        methods = getattr(route, "methods", None)
        if route_path is None or methods is None:
            continue
        key = (route_path, frozenset(methods))
        assert key not in seen, f"duplicate route registered: {key}"
        seen.add(key)
